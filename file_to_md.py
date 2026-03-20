"""
file_to_md.py - 범용 문서 → 옵시디언 마크다운 변환기
지원: Excel(.xlsx/.xls) / Word(.docx) / HWP(.hwp/.hwpx) / PowerPoint(.pptx) / 이미지
"""
import os, re, sys, subprocess, tempfile, shutil, math
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

OBSIDIAN_VAULT_PATH = r"C:\Users\User\Documents\Obsidian Vault"
OUTPUT_FOLDER = os.path.join(OBSIDIAN_VAULT_PATH, "변환문서")
SYSTEM_DIR = os.path.dirname(os.path.abspath(__file__))

LIBREOFFICE_PATHS = [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
]

SUPPORTED_EXT = {
    '.xlsx': 'excel', '.xls': 'excel',
    '.docx': 'word',  '.doc': 'word',
    '.pptx': 'pptx',  '.ppt': 'pptx',
    '.hwp':  'hwp',   '.hwpx': 'hwp',
    '.jpg': 'image', '.jpeg': 'image', '.png': 'image',
    '.gif': 'image', '.webp': 'image', '.bmp': 'image',
}

# ==============================================================================
# 공통 유틸
# ==============================================================================
def get_api_key():
    key = os.environ.get("GEMINI_API_KEY")
    if not key:
        result = subprocess.run(
            ['powershell', '-Command',
             "[System.Environment]::GetEnvironmentVariable('GEMINI_API_KEY', 'User')"],
            capture_output=True, text=True)
        key = result.stdout.strip()
    return key

def get_client():
    from google import genai
    from google.genai import types
    key = get_api_key()
    if not key:
        print("오류: GEMINI_API_KEY 없음"); sys.exit(1)
    return genai.Client(api_key=key), types

def find_libreoffice():
    for p in LIBREOFFICE_PATHS:
        if os.path.exists(p):
            return p
    return None

def gemini_text(client, types, prompt):
    """텍스트 → 마크다운 변환 (Gemini)"""
    import time
    for attempt in range(4):
        try:
            resp = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=prompt,
                config=types.GenerateContentConfig(
                    thinking_config=types.ThinkingConfig(thinking_budget=0)))
            text = resp.text.strip()
            text = re.sub(r'^```\w*\n', '', text)
            text = re.sub(r'\n```$', '', text)
            return text
        except Exception as e:
            if '500' in str(e) or '503' in str(e):
                wait = 10 * (attempt + 1)
                print(f"  서버 오류, {wait}초 후 재시도...", end=" ")
                time.sleep(wait)
            else:
                raise
    return prompt

def gemini_vision(client, types, img_bytes_list, prompt):
    """이미지 + 텍스트 → 마크다운 (Gemini Vision)"""
    import time
    contents = []
    for img in img_bytes_list:
        contents.append(types.Part(
            inline_data=types.Blob(data=img, mime_type="image/png")))
    contents.append(prompt)
    for attempt in range(4):
        try:
            resp = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=contents,
                config=types.GenerateContentConfig(
                    thinking_config=types.ThinkingConfig(thinking_budget=0)))
            text = resp.text.strip()
            text = re.sub(r'^```\w*\n', '', text)
            text = re.sub(r'\n```$', '', text)
            return text
        except Exception as e:
            if '500' in str(e) or '503' in str(e):
                wait = 10 * (attempt + 1)
                print(f"  서버 오류, {wait}초 후 재시도...", end=" ")
                time.sleep(wait)
            else:
                raise
    return "(변환 실패)"

def make_frontmatter(filename, tags=None):
    today = datetime.now().strftime("%Y-%m-%d")
    tag_str = "\n".join(f"  - {t}" for t in (tags or ["변환문서"]))
    return f"---\ntags:\n{tag_str}\n상태: 완료\n날짜: {today}\n원본파일: {filename}\n---\n\n"

# ==============================================================================
# LibreOffice 경유 변환 (PPT, DOC → PDF → 이미지 → Gemini)
# ==============================================================================
def convert_via_libreoffice(path, client, types, out_dir, name, soffice):
    """LibreOffice로 PDF 변환 후 Gemini Vision 처리 (pdf_to_md 파이프라인 재사용)"""
    sys.path.insert(0, SYSTEM_DIR)
    import pdf_to_md as pmd

    with tempfile.TemporaryDirectory() as tmp:
        print(f"  LibreOffice로 PDF 변환 중...", end=" ", flush=True)
        env = os.environ.copy()
        env['HOME'] = tmp  # LibreOffice 사용자 프로파일 임시 경로
        result = subprocess.run(
            [soffice, '--headless', '--norestore', '--nofirststartwizard',
             '--convert-to', 'pdf', path, '--outdir', tmp],
            capture_output=True, text=True, timeout=120, env=env)
        pdf_files = list(Path(tmp).glob("*.pdf"))
        if not pdf_files:
            print(f"\n  변환 실패: {result.stderr[:200]}")
            return None
        pdf_path = str(pdf_files[0])
        total = pmd.get_total_pages(pdf_path)
        print(f"완료 ({total}p)")

        doc_name = input(f"문서명 (엔터='{name}'): ").strip() or name
        part_md = []

        for ps in range(1, total + 1, pmd.CHUNK_PAGES):
            pe = min(ps + pmd.CHUNK_PAGES - 1, total)
            print(f"  [{ps}/{total}] {ps}-{pe}p ...", end=" ", flush=True)
            page_images, _ = pmd.pages_to_images(pdf_path, ps, pe)
            figures = pmd.extract_figures(pdf_path, ps, pe)
            md = pmd.convert_to_md(
                client, types, page_images, figures,
                subject="변환문서", chapter=doc_name,
                page_range=f"{ps}-{pe}",
                first_chunk=(ps == 1),
                page_start=ps,
                img_dir=out_dir,
                name_prefix=name)
            part_md.append(md)
            print("완료")

    return "\n\n---\n\n".join(part_md)

# ==============================================================================
# HWP/HWPX 전용 변환 (3단계 fallback)
# ==============================================================================
def _hwpx_to_text_and_images(path, out_dir, name):
    """HWPX(ZIP+XML) 직접 파싱 → 텍스트 + 이미지 추출"""
    import zipfile, xml.etree.ElementTree as ET
    texts = []
    saved_images = []  # [(zip_name, saved_filename), ...]

    IMG_EXTS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tif', '.tiff'}

    with zipfile.ZipFile(path) as z:
        names = z.namelist()

        # 이미지 추출 (BinData/ 또는 Contents/ 내 이미지 파일)
        img_entries = [n for n in names
                       if Path(n).suffix.lower() in IMG_EXTS
                       or 'bindata' in n.lower()]
        for i, img_entry in enumerate(img_entries, 1):
            ext = Path(img_entry).suffix.lower()
            if ext not in IMG_EXTS:
                continue
            img_data = z.read(img_entry)
            if len(img_data) < 500:  # 너무 작은 파일 제외 (아이콘 등)
                continue
            img_fname = f"{name}_img{i:02d}{ext}"
            img_path = os.path.join(out_dir, img_fname)
            with open(img_path, 'wb') as f:
                f.write(img_data)
            saved_images.append(img_fname)

        # 텍스트 추출 (Contents/*.xml)
        xml_files = sorted([n for n in names
                            if n.startswith('Contents/') and n.endswith('.xml')])
        for xml_file in xml_files:
            try:
                root = ET.fromstring(z.read(xml_file))
                for elem in root.iter():
                    if elem.text and elem.text.strip():
                        texts.append(elem.text.strip())
            except:
                pass

    return '\n'.join(texts), saved_images

def _pyhwp_to_text(path):
    """pyhwp 라이브러리로 HWP 텍스트 추출"""
    import subprocess
    result = subprocess.run(
        ['python', '-m', 'hwp5', 'cat', path],
        capture_output=True, timeout=60)
    if result.returncode == 0:
        return result.stdout.decode('utf-8', errors='ignore')
    # 대안: hwp5txt
    result = subprocess.run(
        ['hwp5txt', path],
        capture_output=True, timeout=60)
    if result.returncode == 0:
        return result.stdout.decode('utf-8', errors='ignore')
    return None

def convert_hwp(path, client, types, out_dir, name, soffice=None):
    """HWP/HWPX → 마크다운 (3단계 fallback)"""
    ext = Path(path).suffix.lower()
    raw_text = None
    saved_images = []

    # 방법 1: HWPX 직접 파싱 (ZIP+XML)
    if ext == '.hwpx':
        print("  HWPX 직접 파싱 중...", end=" ", flush=True)
        try:
            raw_text, saved_images = _hwpx_to_text_and_images(path, out_dir, name)
            if raw_text and len(raw_text) > 50:
                print(f"완료 ({len(raw_text)}자, 이미지 {len(saved_images)}개)")
            else:
                raw_text = None
                print("내용 부족")
        except Exception as e:
            print(f"실패({e})")

    # 방법 2: pyhwp (HWP 바이너리 파싱)
    if not raw_text:
        print("  pyhwp로 텍스트 추출 중...", end=" ", flush=True)
        try:
            raw_text = _pyhwp_to_text(path)
            if raw_text and len(raw_text) > 50:
                print(f"완료 ({len(raw_text)}자)")
            else:
                raw_text = None
                print("실패 또는 내용 부족")
        except Exception as e:
            print(f"실패({e})")

    # 방법 3: LibreOffice → PDF → Gemini Vision
    if not raw_text and soffice:
        print("  LibreOffice 변환 시도 중...")
        return convert_via_libreoffice(path, client, types, out_dir, name, soffice)

    if not raw_text:
        print("\n  모든 변환 방법 실패.")
        print("  해결 방법 중 하나:")
        print("  1. pip install pyhwp")
        print("  2. LibreOffice 설치 (https://www.libreoffice.org/)")
        print("  3. 한글에서 직접 PDF로 저장 후 OCU_PDF변환_실행.bat 사용")
        return None

    # Gemini로 마크다운 변환
    print("  Gemini로 마크다운 변환 중...", end=" ", flush=True)
    prompt = f"""다음은 한글(HWP) 문서에서 추출한 텍스트입니다. 옵시디언 마크다운으로 깔끔하게 변환하세요.
- 제목/소제목 → ##/### 헤더 (문맥으로 판단)
- 목록/번호 → 마크다운 리스트
- 표 형태의 내용 → 마크다운 표
- 중요 용어 **볼드**
- YAML 프론트매터 생략

{raw_text[:8000]}"""
    result = gemini_text(client, types, prompt)
    print("완료")

    # 추출된 이미지를 문서 끝에 첨부
    if saved_images:
        img_section = "\n\n---\n\n## 📎 첨부 이미지\n\n"
        img_section += "\n\n".join(f"![[{fname}]]" for fname in saved_images)
        result += img_section

    return result

# ==============================================================================
# Excel 변환
# ==============================================================================
def convert_excel(path, client, types, out_dir, name):
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl 설치 필요: pip install openpyxl")
        return None

    file_mb = os.path.getsize(path) / (1024 * 1024)
    large_file = file_mb > 20
    if large_file:
        print(f"  대용량 파일 ({file_mb:.1f}MB) → read_only 스트리밍 모드 (이미지 추출 제외)")

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=large_file)
    except PermissionError:
        print("\n  ❌ 파일 접근 불가 (PermissionError)")
        print("  원인: Excel에서 열려 있거나 OneDrive 동기화 중")
        print("  해결: 1) Excel 닫기  2) 파일을 로컬 폴더에 복사 후 재시도")
        return None

    # 시트별 {sheet_name: md_content} 딕셔너리 반환
    sheets = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not any(any(c for c in r) for r in rows):
            print(f"  시트 '{sheet_name}' — 비어있음, 건너뜀")
            continue

        # 실제 사용된 열 범위 계산 (trailing 빈 열 제거)
        used_cols = 0
        for row in rows:
            for j in range(len(row) - 1, -1, -1):
                if row[j] is not None and str(row[j]).strip():
                    used_cols = max(used_cols, j + 1)
                    break
        if used_cols == 0:
            continue

        # 문자열 변환 + trailing 빈 셀 제거
        str_rows = []
        for row in rows:
            cells = [str(c) if c is not None else "" for c in row[:used_cols]]
            # 부족한 열은 빈 문자열로 채움
            cells += [""] * (used_cols - len(cells))
            str_rows.append(cells)

        # 빈 행 제거
        str_rows = [r for r in str_rows if any(c.strip() for c in r)]
        if not str_rows:
            continue

        # 헤더 행 감지
        header_idx = 0
        for i, row in enumerate(str_rows):
            if any(c.strip() for c in row):
                header_idx = i
                break

        def fmt_row(cells):
            return "| " + " | ".join(c.strip() for c in cells) + " |"

        table_lines = []
        for i, row in enumerate(str_rows):
            table_lines.append(fmt_row(row))
            if i == header_idx:
                table_lines.append("| " + " | ".join(["---"] * used_cols) + " |")

        raw = "\n".join(table_lines)

        # 내장 이미지 저장 (read_only 모드에서는 불가)
        img_lines = []
        if large_file:
            pass  # read_only 모드 — 이미지 추출 불가
        for img_obj in getattr(ws, '_images', []) if not large_file else []:
            try:
                img_data = img_obj._data()
                safe_sheet = re.sub(r'[\\/:*?"<>|]', '_', sheet_name)
                fname = f"{name}_{safe_sheet}_{img_obj.anchor._from.row}.png"
                with open(os.path.join(out_dir, fname), 'wb') as f:
                    f.write(img_data)
                img_lines.append(f"\n![[{fname}]]\n")
            except:
                pass

        # 크기에 따라 Gemini 정리 여부 결정 (이미지는 Gemini에 넘기지 않고 나중에 붙임)
        if len(raw) > 6000:
            print(f"  시트 '{sheet_name}' 대용량({len(raw)//1000}KB) → Gemini 생략, 직접 저장")
            result = raw
        else:
            print(f"  시트 '{sheet_name}' Gemini 변환 중...", end=" ", flush=True)
            prompt = f"""다음은 Excel 시트 '{sheet_name}'에서 추출한 데이터입니다. 옵시디언 마크다운으로 변환하세요.
- 표는 마크다운 표 형식 유지
- 합계/소계 행은 **볼드**
- YAML 프론트매터 생략

{raw}"""
            result = gemini_text(client, types, prompt)
            print("완료")

        # 이미지는 Gemini 처리 후 직접 붙임 (Gemini가 삭제하지 못하도록)
        if img_lines:
            result += "\n\n---\n\n## 📎 시트 이미지\n\n" + "\n\n".join(img_lines)

        sheets[sheet_name] = result

    if not sheets:
        print("  내용 없음")
        return None

    return sheets  # dict 반환

# ==============================================================================
# Word(.docx) 변환
# ==============================================================================
def convert_docx(path, client, types, out_dir, name):
    try:
        from docx import Document
        from docx.oxml.ns import qn
    except ImportError:
        print("  python-docx 설치 필요: pip install python-docx")
        return None

    doc = Document(path)

    # rId → (파일명, bytes) 매핑 먼저 구성
    img_map = {}
    img_counter = [0]
    def _get_img_fname(rId):
        if rId in img_map:
            return img_map[rId][0]
        try:
            rel = doc.part.rels[rId]
            if "image" not in rel.reltype:
                return None
            img_bytes = rel.target_part.blob
            ext = rel.target_part.content_type.split('/')[-1].replace('jpeg', 'jpg')
            img_counter[0] += 1
            fname = f"{name}_img{img_counter[0]:02d}.{ext}"
            fpath = os.path.join(out_dir, fname)
            with open(fpath, 'wb') as f:
                f.write(img_bytes)
            img_map[rId] = (fname, img_bytes)
            return fname
        except:
            return None

    parts = []

    for block in doc.element.body:
        tag = block.tag.split('}')[-1]

        if tag == 'p':
            # 이미지가 있는 단락: blip의 r:embed로 rId 추출
            blips = block.findall('.//' + qn('a:blip'))
            if blips:
                for blip in blips:
                    rId = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                    if rId:
                        fname = _get_img_fname(rId)
                        if fname:
                            parts.append(f"\n![[{fname}]]\n")
            else:
                # 텍스트 단락
                para_text = ''.join(
                    t.text or '' for t in block.findall('.//' + qn('w:t')))
                if para_text.strip():
                    parts.append(para_text.strip())

        elif tag == 'tbl':
            rows_text = []
            for tr in block.findall('.//' + qn('w:tr')):
                cells = []
                for tc in tr.findall('.//' + qn('w:tc')):
                    cell_text = ''.join(
                        t.text or '' for t in tc.findall('.//' + qn('w:t')))
                    cells.append(cell_text.strip())
                rows_text.append(cells)
            if rows_text:
                md_rows = []
                for i, row in enumerate(rows_text):
                    md_rows.append("| " + " | ".join(row) + " |")
                    if i == 0:
                        md_rows.append("| " + " | ".join(["---"] * len(row)) + " |")
                parts.append("\n".join(md_rows))

    # 이미지 태그와 텍스트 분리
    img_tags = [p for p in parts if p.strip().startswith('![[')]
    text_parts = [p for p in parts if not p.strip().startswith('![[')]
    img_total = img_counter[0]
    print(f"  이미지 {img_total}개 추출됨")

    raw_content = "\n\n".join(text_parts)
    print("  Gemini로 마크다운 정리 중...", end=" ", flush=True)
    prompt = f"""다음은 Word 문서에서 추출한 텍스트입니다. 옵시디언 마크다운으로 변환하세요.
- 제목/소제목 → ##/### 헤더
- 목록 → - 리스트
- 표 → 마크다운 표 유지
- YAML 프론트매터 생략

{raw_content[:8000]}"""
    result = gemini_text(client, types, prompt)
    print("완료")

    # 이미지는 Gemini 결과 뒤에 직접 붙임
    if img_tags:
        result += "\n\n---\n\n## 📎 첨부 이미지\n\n" + "\n\n".join(img_tags)

    return result

# ==============================================================================
# 이미지 직접 변환
# ==============================================================================
def convert_image(path, client, types, out_dir, name):
    with open(path, 'rb') as f:
        img_bytes = f.read()
    ext = Path(path).suffix.lower().lstrip('.')
    mime = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
            'png': 'image/png', 'gif': 'image/gif',
            'webp': 'image/webp', 'bmp': 'image/bmp'}.get(ext, 'image/png')

    # 이미지 파일 복사
    img_fname = f"{name}{Path(path).suffix}"
    shutil.copy2(path, os.path.join(out_dir, img_fname))

    print("  Gemini Vision 분석 중...", end=" ", flush=True)
    prompt = """이 이미지의 내용을 옵시디언 마크다운으로 정리하세요.
- 텍스트가 있으면 전부 추출
- 표/그래프/도표는 마크다운 표 또는 설명으로 변환
- 중요 내용 볼드 처리
- YAML 프론트매터 생략"""

    contents = [types.Part(inline_data=types.Blob(data=img_bytes, mime_type=mime)), prompt]
    resp = client.models.generate_content(
        model="gemini-2.5-flash", contents=contents,
        config=types.GenerateContentConfig(
            thinking_config=types.ThinkingConfig(thinking_budget=0)))
    text = resp.text.strip()
    text = re.sub(r'^```\w*\n', '', text); text = re.sub(r'\n```$', '', text)
    print("완료")
    return f"![[{img_fname}]]\n\n{text}"

# ==============================================================================
# 메인
# ==============================================================================
def main():
    print("=" * 54)
    print("  범용 문서 → 마크다운 변환기")
    print("  지원: Excel / Word / HWP / PPTX / 이미지")
    print("=" * 54)

    # 파일 경로 입력
    path = input("\n파일 경로를 붙여넣으세요 (드래그 앤 드롭 가능):\n> ").strip().strip('"')
    if not os.path.exists(path):
        print(f"파일 없음: {path}"); return

    ext = Path(path).suffix.lower()
    fmt = SUPPORTED_EXT.get(ext)
    if not fmt:
        print(f"지원하지 않는 형식: {ext}")
        print(f"지원 형식: {', '.join(SUPPORTED_EXT.keys())}"); return

    file_stem = Path(path).stem
    safe_name = re.sub(r'[\\/:*?"<>|]', '_', file_stem)
    today = datetime.now().strftime("%Y-%m-%d")

    # 출력 폴더
    out_dir = os.path.join(OUTPUT_FOLDER, today)
    os.makedirs(out_dir, exist_ok=True)

    print(f"\n형식: {fmt.upper()} | 출력: 변환문서/{today}/")
    print("-" * 54)

    client, types = get_client()
    soffice = find_libreoffice()

    md_body = None

    tags = {"excel": ["변환문서", "Excel"],
            "word":  ["변환문서", "Word"],
            "hwp":   ["변환문서", "HWP"],
            "pptx":  ["변환문서", "PPTX"],
            "image": ["변환문서", "이미지"]}.get(fmt, ["변환문서"])

    saved_files = []

    if fmt == 'excel':
        print("[Excel 변환]")
        sheets = convert_excel(path, client, types, out_dir, safe_name)
        if not sheets:
            print("변환 실패"); return
        # 시트별 개별 파일 저장
        for sheet_name, md_body in sheets.items():
            safe_sheet = re.sub(r'[\\/:*?"<>|]', '_', sheet_name)
            out_filename = f"{safe_name}_{safe_sheet}.md"
            out_path = os.path.join(out_dir, out_filename)
            full_md = make_frontmatter(Path(path).name, tags + [safe_sheet]) + \
                      f"# {sheet_name}\n\n" + md_body
            with open(out_path, 'w', encoding='utf-8') as f:
                f.write(full_md)
            saved_files.append(out_filename)

    else:
        md_body = None

        if fmt == 'word':
            print("[Word 변환]")
            if soffice:
                # LibreOffice → PDF → Gemini Vision (PDF 변환기와 동일 품질)
                print("  LibreOffice → PDF → Gemini Vision 파이프라인")
                md_body = convert_via_libreoffice(path, client, types, out_dir, safe_name, soffice)
            if not md_body:
                # LibreOffice 없거나 실패 시 python-docx fallback
                print("  python-docx 텍스트 추출 모드")
                md_body = convert_docx(path, client, types, out_dir, safe_name)

        elif fmt == 'hwp':
            print("[HWP 변환]")
            md_body = convert_hwp(path, client, types, out_dir, safe_name, soffice)

        elif fmt == 'pptx':
            print("[PPTX 변환]")
            if soffice:
                md_body = convert_via_libreoffice(path, client, types, out_dir, safe_name, soffice)
            else:
                print("  LibreOffice 없음. 설치 필요: https://www.libreoffice.org/")
                return

        elif fmt == 'image':
            print("[이미지 변환]")
            md_body = convert_image(path, client, types, out_dir, safe_name)

        if not md_body:
            print("변환 실패"); return

        out_filename = f"{safe_name}.md"
        out_path = os.path.join(out_dir, out_filename)
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(make_frontmatter(Path(path).name, tags) + md_body)
        saved_files.append(out_filename)

    print("\n" + "=" * 54)
    print(f"✅ 변환 완료 ({len(saved_files)}개 파일)")
    for f in saved_files:
        print(f"   변환문서/{today}/{f}")
    print("=" * 54)

if __name__ == "__main__":
    main()
