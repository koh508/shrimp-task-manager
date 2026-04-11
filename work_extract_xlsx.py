"""
work_extract_xlsx.py
file_manifest.json에서 XLSX(LINELIST / DATA_SHEET)를 읽어 수치 데이터 JSON 추출

출력:
  SYSTEM/work_index/work_extracted/linelist/   → {company_id}_{base}.json
  SYSTEM/work_index/work_extracted/datasheet/  → {company_id}_{base}.json
  SYSTEM/work_index/work_extracted/extract_log.json

API 비용: 0원 (openpyxl 로컬 처리)
실행: python work_extract_xlsx.py [--company 36] [--limit 5] [--force]
"""

import os
import re
import json
import argparse
import hashlib
import logging
from datetime import datetime
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl 미설치. 실행: pip install openpyxl")
    raise

try:
    import xlrd
    _XLRD_OK = True
except ImportError:
    _XLRD_OK = False
    log_tmp = logging.getLogger(__name__)
    log_tmp.warning("xlrd 미설치 — .xls 파일 건너뜀. 설치: pip install xlrd")

# ──────────────────────────────────────────────
# 경로 설정
# ──────────────────────────────────────────────
VAULT_DIR    = r"C:\Users\User\Documents\Obsidian Vault"
WORK_INDEX   = os.path.join(VAULT_DIR, "SYSTEM", "work_index")
MANIFEST_FILE = os.path.join(WORK_INDEX, "file_manifest.json")

EXTRACT_BASE = os.path.join(WORK_INDEX, "work_extracted")
OUT_LINELIST  = os.path.join(EXTRACT_BASE, "linelist")
OUT_DATASHEET = os.path.join(EXTRACT_BASE, "datasheet")
EXTRACT_LOG   = os.path.join(EXTRACT_BASE, "extract_log.json")

for d in [OUT_LINELIST, OUT_DATASHEET]:
    os.makedirs(d, exist_ok=True)

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# 단위 패턴 (셀에서 단위 분리용)
# ──────────────────────────────────────────────
UNIT_PAT = re.compile(
    r"(℃|°[CF]|kPa|MPa|bar(?:a|g)?|kgf/cm2|kg/cm2|"
    r"kg/h|t/h|m3/h|Nm3/h|L/min|m/s|"
    r"kW|MW|kcal/h|Mcal/h|"
    r"mm|\bm\b|inch|\"|%|ppm|pH|"
    r"rpm|Hz|\bkV\b|\bkA\b|\bV\b|\bA\b)"
)

# 숫자값 패턴 (정수·소수·범위)
NUM_PAT = re.compile(r"-?\d[\d,]*\.?\d*(?:\s*~\s*-?\d[\d,]*\.?\d*)?")

# 헤더 후보 키워드 (행이 헤더인지 판별)
HEADER_KEYWORDS = [
    "tag", "item", "service", "description", "parameter",
    "temp", "press", "flow", "온도", "압력", "유량",
    "설계", "정상", "단위", "unit", "value", "remarks",
    "no.", "no ", "번호", "항목",
]

# 무시할 셀값 패턴 (범례, 빈칸 등)
SKIP_PATTERNS = re.compile(
    r"^(n/?a|tbd|tbe|-+|_+|\*+|해당없음|없음|미정|검토중)$",
    re.IGNORECASE
)

# ──────────────────────────────────────────────
# 설비명 정규화 (manifest와 동일한 매핑)
# ──────────────────────────────────────────────
EQUIPMENT_ALIASES = {
    "절탄기":   r"(?i)economi[sz]er|이코노마이저|\bec\b|\becon\b",
    "폐열보일러": r"(?i)hrsg|waste.?heat.?boiler|배열보일러",
    "버너":     r"(?i)burner|점화기",
    "소각로":   r"(?i)incinerator|소각설비|연소로",
    "과열기":   r"(?i)superheater|\bsh\b|s/h",
    "재열기":   r"(?i)reheater|\brh\b|r/h",
    "증기드럼": r"(?i)steam.?drum|\bsd\b|s/d",
    "탈기기":   r"(?i)deaerat|탈기",
    "응축기":   r"(?i)condenser|콘덴서",
    "증발기":   r"(?i)evaporat|증발",
    "펌프":     r"(?i)pump|\bpmp\b",
    "열교환기": r"(?i)heat.?exchanger|\bhex\b|\bh/e\b",
    "압력용기": r"(?i)pressure.?vessel|\bpv\b",
    "탱크":     r"(?i)\btank\b|\btk\b",
    "필터":     r"(?i)filter|strainer|스트레이너",
    "밸브":     r"(?i)\bvalve\b|\bv\b(?=\s*\d)",
    "팬":       r"(?i)\bfan\b|\bblower\b",
    "압축기":   r"(?i)compressor|컴프레서",
}
_EQ_RE = [(name, re.compile(pat)) for name, pat in EQUIPMENT_ALIASES.items()]

def normalize_equipment(text: str) -> str | None:
    """텍스트에서 설비명 추출 (정규화)"""
    if not text:
        return None
    for name, pat in _EQ_RE:
        if pat.search(text):
            return name
    return None

# ──────────────────────────────────────────────
# 헬퍼 함수
# ──────────────────────────────────────────────

def cell_str(cell) -> str:
    """셀값을 문자열로 안전 변환"""
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()

def is_header_row(row_vals: list[str]) -> bool:
    """행이 헤더행인지 판단"""
    joined = " ".join(row_vals).lower()
    hits = sum(1 for kw in HEADER_KEYWORDS if kw in joined)
    return hits >= 2

def extract_value_unit(raw: str) -> tuple[str, str]:
    """'250 ℃' → ('250', '℃')"""
    unit_m = UNIT_PAT.search(raw)
    unit = unit_m.group(0) if unit_m else ""
    # 단위 제거 후 숫자만
    stripped = raw.replace(unit, "").strip() if unit else raw
    num_m = NUM_PAT.search(stripped)
    value = num_m.group(0).replace(",", "") if num_m else stripped
    return value, unit

def chunk_hash(data: dict) -> str:
    s = json.dumps(data, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(s.encode()).hexdigest()[:16]

# ──────────────────────────────────────────────
# 파라미터 컬럼 탐지용 키워드 (라인리스트)
# ──────────────────────────────────────────────
# (헤더명에 아래 키워드 포함 컬럼 → 수치 파라미터 컬럼으로 인식)
PARAM_COL_KEYWORDS = {
    "temp":        ("온도", "℃"),
    "temperature": ("온도", "℃"),
    "온도":        ("온도", "℃"),
    "press":       ("압력", "kgf/cm2G"),
    "pressure":    ("압력", "kgf/cm2G"),
    "압력":        ("압력", "kgf/cm2G"),
    "flow":        ("유량", "kg/h"),
    "flow rate":   ("유량", "kg/h"),
    "유량":        ("유량", "kg/h"),
    "density":     ("밀도", "kg/m3"),
    "밀도":        ("밀도", "kg/m3"),
}

# 태그/서비스 컬럼 후보 키워드
TAG_COL_KEYWORDS    = {"serial no", "line no", "no.", "tag", "태그", "번호"}
SERVICE_COL_KEYWORDS = {"service", "from", "description", "name", "fluid", "항목"}

# ──────────────────────────────────────────────
# 라인리스트 파싱 (2행 합산 헤더 + 파라미터 다중 컬럼)
# ──────────────────────────────────────────────

def _build_combined_header(rows: list, start: int) -> dict[int, str]:
    """
    2행 합산 헤더 처리:
    rows[start] + rows[start+1] 를 컬럼별로 병합
    반환: {col_idx: "combined header string"}
    """
    r1 = [cell_str(c).replace("\n", " ").strip() for c in rows[start]]
    r2 = [cell_str(c).replace("\n", " ").strip() for c in rows[start + 1]] if start + 1 < len(rows) else []

    max_col = max(len(r1), len(r2))
    combined = {}
    for ci in range(max_col):
        v1 = r1[ci] if ci < len(r1) else ""
        v2 = r2[ci] if ci < len(r2) else ""
        merged = f"{v1} {v2}".strip() if (v1 and v2) else (v1 or v2)
        if merged:
            combined[ci] = merged
    return combined


def parse_linelist(ws) -> list[dict]:
    """
    라인리스트 파싱 — 2행 합산 헤더 지원
    각 데이터행 × 파라미터 컬럼 → 레코드 (1행 N건)
    반환: [{tag, service, parameter, value, unit, row_idx}, ...]
    """
    records = []
    rows = list(ws.iter_rows())
    if not rows:
        return records

    # ── 1. 헤더행 탐지 (처음 20행에서 HEADER_KEYWORDS 2개 이상)
    header_row_idx = None
    for ri, row in enumerate(rows[:20]):
        vals = [cell_str(c).replace("\n", " ") for c in row]
        if is_header_row(vals):
            header_row_idx = ri
            break

    if header_row_idx is None:
        log.debug("헤더 미탐지 — 첫 행을 헤더로 가정")
        header_row_idx = 0

    # ── 2. 2행 합산 헤더 빌드
    col_headers = _build_combined_header(rows, header_row_idx)
    data_start = header_row_idx + (2 if header_row_idx + 1 < len(rows) and
                                   not is_header_row([cell_str(c) for c in rows[header_row_idx + 1]])
                                   else 1)

    # ── 3. 컬럼 역할 분류
    tag_cols:   list[int] = []  # 태그/라인번호
    svc_cols:   list[int] = []  # 서비스명
    param_cols: list[tuple[int, str, str]] = []  # (col_idx, param_name, default_unit)

    for ci, hdr in col_headers.items():
        hdr_low = hdr.lower()

        # 파라미터 컬럼 (우선 확인)
        matched_param = False
        for kw, (pname, punit) in PARAM_COL_KEYWORDS.items():
            if kw in hdr_low:
                # 서브타입 판단: oper/normal/design/nor
                sub = ""
                if any(s in hdr_low for s in ["oper", "normal", "nor.", "정상"]):
                    sub = "_정상"
                elif any(s in hdr_low for s in ["design", "설계"]):
                    sub = "_설계"
                param_cols.append((ci, pname + sub, punit))
                matched_param = True
                break

        if matched_param:
            continue

        # 태그 컬럼
        if any(kw in hdr_low for kw in TAG_COL_KEYWORDS):
            tag_cols.append(ci)
            continue

        # 서비스 컬럼
        if any(kw in hdr_low for kw in SERVICE_COL_KEYWORDS):
            svc_cols.append(ci)

    log.debug("헤더행=%d, 데이터시작=%d, 파라미터컬럼=%d개: %s",
              header_row_idx, data_start, len(param_cols),
              [(ci, pn) for ci, pn, _ in param_cols])

    if not param_cols:
        log.debug("파라미터 컬럼 없음 — 단순 1컬럼 파싱으로 폴백")
        return _parse_linelist_simple(rows, header_row_idx, col_headers)

    # ── 4. 데이터 행 파싱
    for ri, row in enumerate(rows[data_start:], start=data_start):
        vals = [cell_str(c) for c in row]
        if not any(vals):
            continue

        # 태그 (첫 번째 태그 컬럼 우선, 없으면 행번호)
        tag = ""
        for tc in tag_cols:
            if tc < len(vals) and vals[tc]:
                tag = vals[tc]
                break

        # 서비스
        service = ""
        for sc in svc_cols:
            if sc < len(vals) and vals[sc]:
                service = vals[sc]
                break

        equipment = normalize_equipment(tag) or normalize_equipment(service)

        # 파라미터별 레코드 생성
        for (pc, pname, default_unit) in param_cols:
            if pc >= len(vals):
                continue
            raw_v = vals[pc]
            if not raw_v or SKIP_PATTERNS.match(raw_v.lower()):
                continue

            value, unit = extract_value_unit(raw_v)
            if not unit:
                unit = default_unit
            if not value:
                continue
            # 숫자인지 확인 (문자열 항목 제거)
            try:
                float(value.replace("~", "").split()[0])
            except (ValueError, IndexError):
                continue

            records.append({
                "tag":       tag,
                "service":   service,
                "parameter": pname,
                "value":     value,
                "unit":      unit,
                "equipment": equipment,
                "row_idx":   ri,
            })

    return records


def _parse_linelist_simple(rows, header_row_idx, col_headers) -> list[dict]:
    """파라미터 컬럼 탐지 실패 시 단순 KEY=VALUE 행 파싱 폴백"""
    records = []
    data_start = header_row_idx + 1

    # 단일 값 컬럼 탐지
    val_col = None
    for ci, hdr in col_headers.items():
        if any(k in hdr.lower() for k in ["value", "값", "수치", "결과"]):
            val_col = ci
            break

    if val_col is None:
        return records

    tag_col = min(col_headers.keys()) if col_headers else 0
    for ri, row in enumerate(rows[data_start:], start=data_start):
        vals = [cell_str(c) for c in row]
        if not any(vals):
            continue
        raw_v = vals[val_col] if val_col < len(vals) else ""
        if not raw_v:
            continue
        value, unit = extract_value_unit(raw_v)
        if not value:
            continue
        tag = vals[tag_col] if tag_col < len(vals) else ""
        records.append({
            "tag": tag, "service": "", "parameter": "값",
            "value": value, "unit": unit, "equipment": None, "row_idx": ri,
        })
    return records

# ──────────────────────────────────────────────
# DATA SHEET 파싱 (KEY-VALUE 세로 형식)
# ──────────────────────────────────────────────

def parse_datasheet(ws) -> list[dict]:
    """
    DATA SHEET는 대부분 2열 구조: [파라미터명 | 값]
    또는 [파라미터명 | 설계값 | 정상값 | 단위]
    반환: [{parameter, value, unit, design_value}, ...]
    """
    records = []
    rows = list(ws.iter_rows())
    if not rows:
        return records

    # 컬럼 너비로 구조 추정 (단순: 2~5열 구조 가정)
    max_col = max(len(row) for row in rows)
    if max_col < 2:
        return records

    # 첫 유효행 찾기 (장비명 행 건너뛰기)
    data_start = 0
    for ri, row in enumerate(rows[:10]):
        vals = [cell_str(c) for c in row]
        non_empty = [v for v in vals if v]
        if len(non_empty) >= 2:
            data_start = ri
            break

    equipment_context = None

    for ri, row in enumerate(rows[data_start:], start=data_start):
        vals = [cell_str(c) for c in row]
        non_empty = [v for v in vals if v]

        if len(non_empty) == 0:
            continue

        # 단일 셀 → 섹션/설비 컨텍스트 업데이트
        if len(non_empty) == 1:
            eq = normalize_equipment(non_empty[0])
            if eq:
                equipment_context = eq
            continue

        # 2열 이상: 첫 열 = 파라미터명, 나머지 = 값들
        param_raw = vals[0] if vals else ""
        if not param_raw or SKIP_PATTERNS.match(param_raw.lower()):
            continue

        # 값/단위 열 추출
        value_cells = [v for v in vals[1:] if v]
        if not value_cells:
            continue

        # 단위 분리: 마지막 셀이 단위처럼 보이면 분리
        unit = ""
        if len(value_cells) >= 2 and UNIT_PAT.match(value_cells[-1]):
            unit = value_cells[-1]
            value_cells = value_cells[:-1]

        # 설계값 / 정상값 구분 (열이 3개 이상이면)
        design_value = ""
        if len(value_cells) >= 2:
            design_value = value_cells[0]
            raw_v = value_cells[1]
        else:
            raw_v = value_cells[0]

        v, u = extract_value_unit(raw_v)
        if not u and unit:
            u = unit
        if not v:
            v = raw_v  # 숫자 아니어도 텍스트 값은 보존

        equipment = normalize_equipment(param_raw) or equipment_context

        records.append({
            "parameter":    param_raw,
            "value":        v,
            "unit":         u,
            "design_value": design_value,
            "equipment":    equipment,
            "row_idx":      ri,
        })

    return records

# ──────────────────────────────────────────────
# 단일 XLSX 파일 처리
# ──────────────────────────────────────────────

def process_xlsx(manifest_entry: dict) -> dict:
    """
    manifest 항목 1개 처리 → 추출 결과 dict 반환
    {ok: bool, records: [...], sheets: int, error: str}
    """
    fpath = manifest_entry["abs_path"]
    doc_type = manifest_entry.get("doc_type", "")
    company_id = manifest_entry.get("company_id", "00")
    equipment = manifest_entry.get("equipment", "")
    rev = manifest_entry.get("rev", "")
    base = Path(fpath).stem[:40]  # 파일명 길이 제한

    result = {
        "source_path": fpath,
        "company_id":  company_id,
        "equipment":   equipment,
        "doc_type":    doc_type,
        "rev":         rev,
        "is_latest":   manifest_entry.get("is_latest", True),
        "records":     [],
        "sheets_processed": 0,
        "ok":    False,
        "error": "",
    }

    is_xls = Path(fpath).suffix.lower() == ".xls"

    if is_xls:
        if not _XLRD_OK:
            result["error"] = "xlrd_missing"
            return result
        return _process_xls(fpath, result, doc_type, equipment)

    try:
        wb = load_workbook(fpath, read_only=True, data_only=True)
    except Exception as e:
        result["error"] = f"open_fail: {e}"
        log.warning("열기 실패 %s: %s", fpath, e)
        return result

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
        except Exception:
            continue

        # 빈 시트 건너뜀 (max_row None 이면 빈 시트)
        if ws.max_row is None or ws.max_row < 2:
            continue

        try:
            if doc_type == "LINELIST":
                recs = parse_linelist(ws)
            else:  # DATA_SHEET
                recs = parse_datasheet(ws)
        except Exception as e:
            log.warning("시트 파싱 오류 %s[%s]: %s", fpath, sheet_name, e)
            continue

        # 설비명 보강: manifest의 equipment로 fallback
        for r in recs:
            if not r.get("equipment") and equipment:
                r["equipment"] = equipment
            r["sheet"] = sheet_name
            r["chunk_hash"] = chunk_hash({
                "path": fpath, "sheet": sheet_name,
                "row": r.get("row_idx", 0), "value": r.get("value", "")
            })

        result["records"].extend(recs)
        result["sheets_processed"] += 1

    wb.close()
    result["ok"] = True
    return result

# ──────────────────────────────────────────────
# 결과 저장
# ──────────────────────────────────────────────

def _xls_row_to_strs(sheet, ri: int) -> list[str]:
    """xlrd 시트 한 행을 문자열 리스트로 변환"""
    vals = []
    for ci in range(sheet.ncols):
        cell = sheet.cell(ri, ci)
        if cell.ctype == 0:   # empty
            vals.append("")
        elif cell.ctype == 1: # text
            vals.append(str(cell.value).strip())
        elif cell.ctype == 2: # number
            v = cell.value
            vals.append(str(int(v)) if v == int(v) else str(v))
        elif cell.ctype == 3: # date — skip
            vals.append("")
        else:
            vals.append(str(cell.value).strip())
    return vals


class _XlsWsAdapter:
    """xlrd Sheet → parse_linelist/parse_datasheet 호환 래퍼"""
    def __init__(self, sheet):
        self._s = sheet
        self.max_row = sheet.nrows
        self.max_column = sheet.ncols

    def iter_rows(self, max_row=None):
        nrows = min(self._s.nrows, max_row) if max_row else self._s.nrows
        for ri in range(nrows):
            yield [_XlsCellAdapter(self._s.cell(ri, ci)) for ci in range(self._s.ncols)]


class _XlsCellAdapter:
    """xlrd Cell → openpyxl cell 인터페이스 호환"""
    def __init__(self, cell):
        if cell.ctype == 0:
            self.value = None
        elif cell.ctype == 2:
            v = cell.value
            self.value = int(v) if v == int(v) else v
        elif cell.ctype == 3:
            self.value = None  # date — skip
        else:
            self.value = str(cell.value).strip() or None


def _process_xls(fpath: str, result: dict, doc_type: str, equipment: str) -> dict:
    """xlrd로 .xls 파일 처리 — parse_linelist/parse_datasheet 재사용"""
    try:
        wb = xlrd.open_workbook(fpath)
    except Exception as e:
        result["error"] = f"xls_open_fail: {e}"
        log.warning("XLS 열기 실패 %s: %s", fpath, e)
        return result

    for sheet_name in wb.sheet_names():
        try:
            raw_sheet = wb.sheet_by_name(sheet_name)
        except Exception:
            continue

        if raw_sheet.nrows < 2:
            continue

        ws = _XlsWsAdapter(raw_sheet)

        try:
            if doc_type == "LINELIST":
                recs = parse_linelist(ws)
            else:
                recs = parse_datasheet(ws)
        except Exception as e:
            log.warning("XLS 시트 파싱 오류 %s[%s]: %s", fpath, sheet_name, e)
            continue

        for r in recs:
            if not r.get("equipment") and equipment:
                r["equipment"] = equipment
            r["sheet"] = sheet_name
            r["chunk_hash"] = chunk_hash({
                "path": fpath, "sheet": sheet_name,
                "row": r.get("row_idx", 0), "value": r.get("value", "")
            })

        result["records"].extend(recs)
        result["sheets_processed"] += 1

    result["ok"] = True
    return result


def save_result(result: dict):
    """처리 결과를 doc_type에 따라 JSON으로 저장"""
    company_id = result["company_id"]
    base = Path(result["source_path"]).stem[:40]
    # 파일명 안전화
    safe_base = re.sub(r"[^\w가-힣\-]", "_", base)
    fname = f"{company_id}_{safe_base}.json"

    if result["doc_type"] == "LINELIST":
        out_path = os.path.join(OUT_LINELIST, fname)
    else:
        out_path = os.path.join(OUT_DATASHEET, fname)

    payload = {
        "meta": {
            "source_path": result["source_path"],
            "company_id":  result["company_id"],
            "equipment":   result["equipment"],
            "doc_type":    result["doc_type"],
            "rev":         result["rev"],
            "is_latest":   result["is_latest"],
            "extracted_at": datetime.now().isoformat(timespec="seconds"),
            "record_count": len(result["records"]),
        },
        "records": result["records"],
    }

    tmp = out_path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    os.replace(tmp, out_path)
    return out_path

# ──────────────────────────────────────────────
# 메인
# ──────────────────────────────────────────────

def load_manifest() -> list[dict]:
    if not os.path.exists(MANIFEST_FILE):
        log.error("manifest 없음: %s — work_make_manifest.py를 먼저 실행하세요", MANIFEST_FILE)
        raise FileNotFoundError(MANIFEST_FILE)
    with open(MANIFEST_FILE, encoding="utf-8") as f:
        return json.load(f)

def main():
    parser = argparse.ArgumentParser(description="XLSX 수치 추출기")
    parser.add_argument("--company", type=str, default=None, help="특정 업체 ID만 처리 (예: 36)")
    parser.add_argument("--limit",   type=int, default=None, help="처리 개수 제한 (테스트용)")
    parser.add_argument("--force",   action="store_true",    help="이미 추출된 파일도 재처리")
    parser.add_argument("--latest-only", action="store_true", default=True,
                        help="최신 Rev 파일만 처리 (기본값: on)")
    args = parser.parse_args()

    manifest = load_manifest()

    # 필터링: XLSX + LINELIST or DATA_SHEET
    targets = [
        e for e in manifest
        if Path(e["abs_path"]).suffix.lower() in {".xlsx", ".xls"}
        and not Path(e["abs_path"]).name.startswith("~$")   # Office 잠금파일 제외
        and e.get("doc_type") in {"LINELIST", "DATA_SHEET"}
        and (not args.latest_only or e.get("is_latest", True))
        and (args.company is None or str(e.get("company_id", "")) == args.company)
    ]

    log.info("처리 대상: %d개 파일 (전체 manifest: %d개)", len(targets), len(manifest))

    if args.limit:
        targets = targets[:args.limit]
        log.info("--limit 적용: %d개만 처리", len(targets))

    # 기존 추출 파일 목록 (force 아닐 때 skip 판단용)
    existing = set()
    if not args.force:
        for d in [OUT_LINELIST, OUT_DATASHEET]:
            for f in os.listdir(d):
                if f.endswith(".json"):
                    existing.add(f)

    stats = {"processed": 0, "skipped": 0, "failed": 0, "total_records": 0}
    log_entries = []

    for entry in targets:
        company_id = entry.get("company_id", "00")
        base = Path(entry["abs_path"]).stem[:40]
        safe_base = re.sub(r"[^\w가-힣\-]", "_", base)
        expected_fname = f"{company_id}_{safe_base}.json"

        if not args.force and expected_fname in existing:
            stats["skipped"] += 1
            continue

        log.info("[%s] %s", entry.get("doc_type"), Path(entry["abs_path"]).name)
        result = process_xlsx(entry)

        if not result["ok"]:
            stats["failed"] += 1
            log_entries.append({"path": entry["abs_path"], "ok": False, "error": result["error"]})
            continue

        if result["records"]:
            out_path = save_result(result)
            log.info("  → %d건 저장: %s", len(result["records"]), Path(out_path).name)
        else:
            log.info("  → 추출된 레코드 없음 (빈 시트 또는 파싱 미적중)")

        stats["processed"] += 1
        stats["total_records"] += len(result["records"])
        log_entries.append({
            "path":    entry["abs_path"],
            "ok":      True,
            "doc_type": result["doc_type"],
            "records": len(result["records"]),
            "sheets":  result["sheets_processed"],
        })

    # 로그 저장
    log_payload = {
        "run_at": datetime.now().isoformat(timespec="seconds"),
        "args": vars(args),
        "stats": stats,
        "entries": log_entries,
    }
    tmp = EXTRACT_LOG + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(log_payload, f, ensure_ascii=False, indent=2)
    os.replace(tmp, EXTRACT_LOG)

    print(f"\n완료: 처리 {stats['processed']}개 / 스킵 {stats['skipped']}개 / 실패 {stats['failed']}개")
    print(f"총 추출 레코드: {stats['total_records']}건")
    print(f"로그: {EXTRACT_LOG}")

if __name__ == "__main__":
    main()
